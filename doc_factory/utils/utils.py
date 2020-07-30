
from PIL import Image
from tesserocr import PyTessBaseAPI
from pdf2image import convert_from_path, convert_from_bytes
from pdf2image.exceptions import (
    PDFInfoNotInstalledError,
    PDFPageCountError,
    PDFSyntaxError
)
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from xlrd import open_workbook
from xlutils.copy import copy
from xlwt import XFStyle

ex = open_workbook('Exmpl.xls', formatting_info=1)
wb = copy(ex)
wb.save('test_2.xls')
# exsh = ex.sheet_by_index(0)
# testcell = exsh.cell_value(6, 0)
#
# wb = copy(ex)
#
# wbsh = wb.get_sheet(0)
# style = XFStyle()
# style.alignment.wrap = 1
# wbsh.write(6, 0, testcell, style)
# wb.save('test_1.xls')


# images = convert_from_path('pdf/2-38-19-2788.pdf')
# n = 1
# for img in images:
#     img.save('pdf/new_{}.png'.format(n), dpi=(300, 300))
#     n += 1
# #
#
# images_for_ocr = ['pdf/new_1.png']
#
# with PyTessBaseAPI(lang='rus') as api:
#     for im in images_for_ocr:
#         api.SetImageFile(im)
#         str = api.GetUTF8Text()
# #
# # start = str.find('1. Наименование объекта', )
# # end = str.find('2.')
# # title = str[start + 25 : end].strip()
# wb = Workbook()
# ws1 = wb.active
